#new_nxthink
from selenium  import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.chrome.service import Service

from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from datetime import date
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import StaleElementReferenceException
from datetime import date, timedelta
import tkinter as tk
from tkinter import messagebox
import shutil
from datetime import datetime

today = datetime.now()
current_hour = today.strftime("%H-%M")  # Using dashes instead of colons
formatted_date = today.strftime(f"{current_hour} %d-%m-%Y")
source_file = r'\\10.199.199.35\soc team\SOC_Daily Report\2023\NXThink\Nexthink_new.xlsx'
destination_file = r"\\10.199.199.35\soc team\Abdelrahman Ataa\Backup/" + formatted_date + ".xlsx"
shutil.copyfile(source_file, destination_file)

class InstaBot:
    def __init__(self,username,password):  
        #webdriver
        try:
            serv_obj = service= Service(r'\\10.199.199.35\soc team\Abdelrahman Ataa\Nxthink\dist\geckodriver.exe')
            ops=webdriver.FirefoxOptions()
            self.driver = webdriver.Firefox(service=serv_obj,options=ops)
            #launch
            self.driver.get("https://nxportal.egyptpost.local")
            self.driver.implicitly_wait(10)
        except Exception as e :
            messagebox.showerror("Error", f"There is problem with geckodriver.exe or problem with internet \n  geckodriver هنالك مشكلة بالانترنت او بملف ")
            #try to use chromedriver instead
            try:
                serv_obj = service= Service(r'\\10.199.199.35\soc team\Abdelrahman Ataa\Nxthink\dist\chromedriver.exe')
                ops=webdriver.ChromeOptions()
                self.driver = webdriver.Chrome(service=serv_obj,options=ops)
                #launch
                self.driver.get("https://nxportal.egyptpost.local")
                self.driver.implicitly_wait(15)
            except Exception as e :
                messagebox.showerror("Error", f"There is problem with chromedriver.exe or problem with internet \n  chromedriver هنالك مشكلة بالانترنت او بملف ")
                return
        try:
            self.username= username
            self.password= password
            #username
            self.driver.find_element(By.XPATH, "//input[@id='LS_username_input']").send_keys(username)
            #password
            self.driver.find_element(By.XPATH, "//input[@id='LS_password_input']").send_keys(password)
            #login
            self.driver.find_element(By.XPATH, "//button[@id='LS_login_button']").click()
            self.driver.maximize_window()
            sleep(10)
        except Exception as e :
            messagebox.showerror("Error","Username or password uncorrecty or can't reach the page \n هنالك مشكلة بالانترنت او ان الاميل أوالباسورد خطأ")
            return



    def get_tickets_data(self):
        self.driver.find_element(By.XPATH, "//div[@id='TNV_menu_toggler']").click()
        sleep(1)
        self.driver.find_element(By.XPATH, "//label[@for='TNV_hhistory_radio']").click()
        sleep(1)
        self.driver.find_element(By.XPATH, "//button[@id='TNV_ok_button']").click()
        sleep(1)
        item = self.driver.find_element(By.XPATH, "//div[@id='TNV_backward_img']")
        for i in range(1,25):
            item.click()
        #gg
        for i in range(1,25):
            current_time =  self.driver.find_element(By.XPATH, "//span[@id='TNV_date_text']").text
            current_day = self.driver.find_element(By.XPATH, "//span[@id='TNV_scope_text']").text

            if current_day == "Yesterday":
                yesterday = date.today() - timedelta(days=1)
                result = yesterday.strftime("%d-%m-%Y")
            elif current_day == "Today":
                today = date.today()
                result = today.strftime("%d-%m-%Y")   
            ser = 2
            for ser in range(ser,8):
                #menu tap
                try:
                    self.driver.find_element(By.XPATH, "//button[@id='MNV_dashboards_button']").click()
                    sleep(2)
                    self.driver.find_element(By.XPATH, "//ul[@aria-label='Egypt Post Main Services']//li["+str(ser)+"]").click()
                except Exception as e:
                    messagebox.showerror("Error","there's problem with menu page and can't reach  \n هنالك مشكلة بالصفحة الرئيسية ولا يستطع الوصول الي صفحة الانتقالات")
                    return
                row = 2
                for row in range(row,12):
                    try:
                        Devices_issues= self.driver.find_element(By.XPATH, "//body[1]/div[1]/div[1]/div[2]/div[3]/div[1]/div[1]/div[1]/div[6]/div[6]/div[2]/div[1]/div[2]/div[1]/div[1]/div[3]/div[4]/div[1]/div[1]/table[1]/tbody[1]/tr["+str(row)+"]/td[2]")
                        print(Devices_issues.text)
                    except Exception as e:
                        messagebox.showerror("Error","theres problem with getting devices issues\n لا يستطيع الوصول الي المشاكل الموجودة ")
                        return
                    if Devices_issues.text == '0':
                        pass 
                    else:
                        try:
                            sheet_path = r'\\10.199.199.35\soc team\SOC_Daily Report\2023\NXThink\Nexthink_new.xlsx'
                            wb = openpyxl.load_workbook(sheet_path)
                            sheet = wb['Sheet1']   

                            max_rows = sheet.max_row+1
                            #servers_name----------------------------
                            Service_name =  self.driver.find_element(By.XPATH, "//span[@id='DNV_dashboard_span']")
                            cell = sheet.cell(max_rows, 1)
                            cell.value = Service_name.text
                            # Center the text
                            alignment = Alignment(horizontal='center', vertical='center')
                            cell.alignment = alignment

                            # Apply borders and set color to gray
                            border = Border(
                                left=Side(border_style='thin', color='FF808080'),
                                right=Side(border_style='thin', color='FF808080'),
                                top=Side(border_style='thin', color='FF808080'),
                                bottom=Side(border_style='thin', color='FF808080')
                            )
                            cell.border = border
                            #Date----------------------
                            cell = sheet.cell(max_rows, 3)
                            cell.value = result
                            cell.border = border
                            cell.alignment = alignment

                            #area--------------------------
                            area = self.driver.find_element(By.XPATH, "//body[1]/div[1]/div[1]/div[2]/div[3]/div[1]/div[1]/div[1]/div[6]/div[6]/div[2]/div[1]/div[2]/div[1]/div[1]/div[3]/div[4]/div[1]/div[1]/table[1]/tbody[1]/tr["+str(row)+"]/td[1]")
                            cell = sheet.cell(max_rows, 4)
                            cell.value = area.text
                            cell.border = border
                            cell.alignment = alignment
                            #time---------------------------
                            cell = sheet.cell(max_rows, 2)
                            cell.value =  current_time
                            cell.border = border
                            cell.alignment = alignment
                            #issues
                            cell = sheet.cell(max_rows, 5)
                            cell.value =  Devices_issues.text
                            cell.border = border
                            cell.alignment = alignment
                            try:
                                wb.save(sheet_path)
                            except FileNotFoundError as e:
                                messagebox.showerror("Error","File not found\n الفايل راح فين ؟ بطلو لعب ")
                                return
                            except PermissionError as e:
                                messagebox.showerror("Error","Permission denied\n لا يوجد صلاحية للتعديل يمكن ان يكون الشيت مفتوح ")
                            except ValueError as e:
                                messagebox.showerror("Error","Invalid value")
                            except Exception as e:
                                messagebox.showerror("Error","فيه مشكلة بالشيت ياسطا")
                                return
                        except NoSuchElementException as e:
                            messagebox.showerror("Error","Element not found\n لا يستطيع الوصول لعنصر بالموقع")
                        except StaleElementReferenceException as e:
                            messagebox.showerror("Error","Element is no longer attached to the DOM")
                        except WebDriverException as e:
                            messagebox.showerror("Error","An unexpected error occurred during the operation")
                            return
            self.driver.find_element(By.XPATH, "//div[@id='TNV_forward_img']").click()

    def Quit(self):
        self.driver.quit()






mybot=InstaBot("w_abdelrahman.ataa", "a1591997A!")
mybot.get_tickets_data()
mybot.Quit()



