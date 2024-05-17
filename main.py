#new_nxthink
from selenium  import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.chrome.service import Service
import os
from dotenv import load_dotenv
from time import sleep
env_path = "C:\\env\\.env"  
load_dotenv(dotenv_path=env_path)
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from datetime import date
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import StaleElementReferenceException
from datetime import date, timedelta
from datetime import datetime
import os 
today = datetime.now()
current_hour = today.strftime("%H-%M") 
formatted_date = today.strftime(f"{current_hour} %d-%m-%Y")
source_file = os.getenv("source_file")
destination_file = f'{os.getenv("destination_file")}\{formatted_date}.xlsx"'

class InstaBot:
    def __init__(self,username,password):  
        #webdriver
        try:
            serv_obj = service= Service(os.getenv("geckodriver"))
            ops=webdriver.FirefoxOptions()
            self.driver = webdriver.Firefox(service=serv_obj,options=ops)
            #launch
            self.driver.get("https://nxportal.egyptpost.local")
            self.driver.implicitly_wait(10)
        except Exception as e :
            print( f"There is problem with geckodriver.exe or problem with internet \n  geckodriver هنالك مشكلة بالانترنت او بملف ")
            #try to use chromedriver instead
            try:
                serv_obj = service= Service(os.getenv("chromedriver"))
                ops=webdriver.ChromeOptions()
                self.driver = webdriver.Chrome(service=serv_obj,options=ops)
                #launch
                self.driver.get("https://nxportal.egyptpost.local")
                self.driver.implicitly_wait(15)
            except Exception as e :
                print( f"There is problem with chromedriver.exe or problem with internet \n  chromedriver هنالك مشكلة بالانترنت او بملف ")
                return
        try:
            self.username= username
            sleep(1)
            self.password= password
            sleep(1)
            #username
            self.driver.find_element(By.XPATH, "//input[@id='LS_username_input']").send_keys(username)
            #password
            self.driver.find_element(By.XPATH, "//input[@id='LS_password_input']").send_keys(password)
            #login
            self.driver.find_element(By.XPATH, "//button[@id='LS_login_button']").click()
            self.driver.maximize_window()
            sleep(10)
        except Exception as e :
            print("Username or password uncorrecty or can't reach the page \n هنالك مشكلة بالانترنت او ان الاميل أوالباسورد خطأ")
            return



    def get_tickets_data(self):
        self.driver.find_element(By.XPATH, "//div[@id='TNV_menu_toggler']").click()
        sleep(1)
        self.driver.find_element(By.XPATH, "//label[@for='TNV_hhistory_radio']").click()
        sleep(1)
        self.driver.find_element(By.XPATH, "//button[@id='TNV_ok_button']").click()
        sleep(1)
        item = self.driver.find_element(By.XPATH, "//div[@id='TNV_backward_img']")
        for i in range(1,13):
            item.click()
        #gg
        for i in range(1,13):
            current_time =  self.driver.find_element(By.XPATH, "//span[@id='TNV_date_text']").text
            current_day = self.driver.find_element(By.XPATH, "//span[@id='TNV_scope_text']").text

            if current_day == "Yesterday":
                yesterday = date.today() - timedelta(days=1)
                result = yesterday.strftime("%d-%m-%Y")
            elif current_day == "Today":
                today = date.today()
                result = today.strftime("%d-%m-%Y")   
            ser = 2
            for ser in range(ser,8):
                #menu tap
                try:
                    self.driver.find_element(By.XPATH, "//button[@id='MNV_dashboards_button']").click()
                    sleep(2)
                    self.driver.find_element(By.XPATH, "//ul[@aria-label='Egypt Post Main Services']//li["+str(ser)+"]").click()
                except Exception as e:
                    print("there's problem with menu page and can't reach  \n هنالك مشكلة بالصفحة الرئيسية ولا يستطع الوصول الي صفحة الانتقالات")
                    return
                row = 2
                for row in range(row,12):
                    try:
                        Devices_issues= self.driver.find_element(By.XPATH, "//body[1]/div[1]/div[1]/div[2]/div[3]/div[1]/div[1]/div[1]/div[6]/div[6]/div[2]/div[1]/div[2]/div[1]/div[1]/div[3]/div[4]/div[1]/div[1]/table[1]/tbody[1]/tr["+str(row)+"]/td[2]")
                        print(Devices_issues.text)
                    except Exception as e:
                        print("theres problem with getting devices issues\n لا يستطيع الوصول الي المشاكل الموجودة ")
                        return
                    if Devices_issues.text == '0':
                        pass 
                    else:
                        try:
                            sheet_path = os.getenv("sheet_path")
                            wb = openpyxl.load_workbook(sheet_path)
                            sheet = wb['Sheet1']   

                            max_rows = sheet.max_row+1
                            #servers_name----------------------------
                            Service_name =  self.driver.find_element(By.XPATH, "//span[@id='DNV_dashboard_span']")
                            cell = sheet.cell(max_rows, 1)
                            cell.value = Service_name.text
                            # Center the text
                            alignment = Alignment(horizontal='center', vertical='center')
                            cell.alignment = alignment

                            # Apply borders and set color to gray
                            border = Border(
                                left=Side(border_style='thin', color='FF808080'),
                                right=Side(border_style='thin', color='FF808080'),
                                top=Side(border_style='thin', color='FF808080'),
                                bottom=Side(border_style='thin', color='FF808080')
                            )
                            cell.border = border
                            #Date----------------------
                            cell = sheet.cell(max_rows, 3)
                            cell.value = result
                            cell.border = border
                            cell.alignment = alignment

                            #area--------------------------
                            area = self.driver.find_element(By.XPATH, "//body[1]/div[1]/div[1]/div[2]/div[3]/div[1]/div[1]/div[1]/div[6]/div[6]/div[2]/div[1]/div[2]/div[1]/div[1]/div[3]/div[4]/div[1]/div[1]/table[1]/tbody[1]/tr["+str(row)+"]/td[1]")
                            cell = sheet.cell(max_rows, 4)
                            cell.value = area.text
                            cell.border = border
                            cell.alignment = alignment
                            #time---------------------------
                            cell = sheet.cell(max_rows, 2)
                            cell.value =  current_time
                            cell.border = border
                            cell.alignment = alignment
                            #issues
                            cell = sheet.cell(max_rows, 5)
                            cell.value =  Devices_issues.text
                            cell.border = border
                            cell.alignment = alignment
                            try:
                                wb.save(sheet_path)
                            except FileNotFoundError as e:
                                print("File not found\n الفايل راح فين ؟ بطلو لعب ")
                                return
                            except PermissionError as e:
                                print("Permission denied\n لا يوجد صلاحية للتعديل يمكن ان يكون الشيت مفتوح ")
                            except ValueError as e:
                                print("Invalid value")
                            except Exception as e:
                                print("فيه مشكلة بالشيت ياسطا")
                                return
                        except NoSuchElementException as e:
                            print("Element not found\n لا يستطيع الوصول لعنصر بالموقع")
                        except StaleElementReferenceException as e:
                            print("Element is no longer attached to the DOM")
                        except WebDriverException as e:
                            print("An unexpected error occurred during the operation")
                            return
            self.driver.find_element(By.XPATH, "//div[@id='TNV_forward_img']").click()

    def Quit(self):
        self.driver.quit()




username = os.getenv("name")
passwd =  os.getenv("passwd")
print(f"{username}: {passwd}")
mybot=InstaBot(username, passwd)
mybot.get_tickets_data()
mybot.Quit()




